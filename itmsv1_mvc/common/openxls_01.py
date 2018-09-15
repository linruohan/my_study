from openpyxl.workbook import Workbook
from __future__ import absolute_import
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
#新建一个workbook
wb = Workbook()
#第一个sheet是ws
ws = wb.worksheets[0]
#设置ws的名称
ws.title = u"下单统计"
#给A1赋值
ws.cell('A1').value = '%s'%("跟随总数")
#给A2赋值
#先把数字转换成字母
col = get_column_letter(1)
#赋值
ws.cell('%s%s'%(col, 2)).value = '%s' % ("A2")
#字体修改样式
##颜色
ws.cell('A2').style.font.color.index =Color.GREEN
##字体名称
ws.cell('A2').style.font.name ='Arial'
##字号
ws.cell('A2').style.font.size =8
##加粗
ws.cell('A2').style.font.bold =True
##不知道干啥用的
ws.cell('A2').style.alignment.wrap_text =True
##背景 好像不太好用 是个BUG
ws.cell('A2').style.fill.fill_type =Fill.FILL_SOLID
ws.cell('A2').style.fill.start_color.index =Color.DARKRED
##修改某一列宽度
ws.column_dimensions["C"].width =60.0
##增加一个表
ws = wb.create_sheet(title = u'结单统计')
##保存生成xml
file_name = 'test.xlsx'
file_dir = '/home/x/'
dest_filename = '%s%s'%(file_dir,file_name)
ew = ExcelWriter(workbook = wb)
ew = ExcelWriter(workbook = wb)