from operator import itemgetter
from datetime import datetime, timedelta
from xlrd import open_workbook, cellname, xldate_as_tuple, \
    XL_CELL_NUMBER, XL_CELL_DATE, XL_CELL_TEXT, XL_CELL_BOOLEAN, \
    XL_CELL_ERROR, XL_CELL_BLANK, XL_CELL_EMPTY, error_text_from_code
from xlwt import easyxf, Workbook
from xlutils.copy import copy as copy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from itmsv1_mvc.common.mylog import Log as log
import pandas as pd
import re
log = log ()
''' df[col]：根据列名，并以Series的形式返回列 
    df[[col1, col2]]：以DataFrame形式返回多列 
        s.iloc[0]：按位置选取数据 
        s.loc[‘index_one’]：按索引选取数据 
    df.iloc[0,:]：返回第一行 
    df.iloc[0,0]：返回第一列的第一个元素 
    # pd.concat ([self.df, df1]).to_excel (self.path, sheet_name=self.page)'''


class ExcelUtil ():
    '''excel文件的操作
    根据【页面元素】的名字获取其值，并进行相应的读写操作'''

    def __init__(self, path, sheet_str):
        self.path = path
        self.sheet_str = sheet_str
        self.sheetname=None
        self.workbook = open_workbook (path, formatting_info=False, on_demand=True)

    def get_sheet(self):
        try:
            if not isinstance (self.sheet_str, str):
                log.error ('s_name为字符串，不能为其他信息')
            self.sheetname = [s for s in self.workbook.sheet_names () if self.sheet_str in s][0]
            log.info('现在读取的是%s中的%s【sheet】'%(self.path,self.sheetname))
        except Exception  as e:
            log.error ('s_name[%s]错误或者不存在！！！%s' % (self.sheet_str, e))
        return self.sheetname
    def read_df(self):
        self.sheetname=self.get_sheet()
        self.df = pd.read_excel (self.path,self.sheetname , index_col=None,na_values='')
        print(self.df)
        # print (self.df['关键字'][0])
        return self.df

    def get_col_info(self, key):
        df = self.df
        try:
            log.info ('【%s】列信息显示如下：\n%s' % (key, df[key]))
        except KeyError:
            log.error ('【%s】不是列名，请重新查看' % key)
        return df[key]
    def get_col_index(self,key):
        column = [list(df.columns).index(v) for v in list(df.columns) if key in v]
        try:
            print(column)
            print ('根据key【%s】获取列号为：%s' % (key, column[0]))
        except Exception as e:
            log.error ('列中无该值存在！%s'%e)
        return column[0]
    def get_row_index(self,key):
        row = [i for i in range(len(list(df.values)))
            for v in list(df.values)[i]
                if key in str(v)]
        try:
            print (row)
            print ('根据key【%s】获取行号为：%s' % (key, row[0]))
        except Exception as e:
            log.error ('行中无该值存在！')
        return row[0]
    def get_element_row_num(self, key):
        df = self.df
        key_pattern = r'^%s$' % key
        row = []
        try:
            column = [i for i in df.columns if '页面元素' in i][0]
            row = [k for k, v in df[column].items () if re.findall (key_pattern, v)][0]
            log.info ('第【%s】行' % row)
        except IndexError:
            log.error ('错误，该行不存在，%s' % df.columns)
        return row

    def get_row_info(self, key):
        df = self.df
        num=self.get_row_index (key)
        row_info = df.iloc[num,:]
        log.info ('【%s】行信息显示如下：\n%s' % (key, row_info))
        return row_info

    def get_selector(self, key):
        df = self.df
        row_info = self.get_row_info (key)
        by = row_info['by']
        bystyle = row_info['by_style']
        xpath = by + '=>' + bystyle
        log.info ('【%s】元素的selector信息显示如下：\n%s' % (key, xpath))
        return xpath[0]
    def save(self,df):
        book = load_workbook (self.path)
        writer = pd.ExcelWriter (self.path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict ((ws.title, ws) for ws in book.worksheets)
        df.to_excel (writer, self.sheetname, index=False)
        writer.save ()
        log.info ('保存成功！')
    def set_value(self, element,col_value, value):
        df = self.df
        row = self.get_row_index(element)
        print(row)
        col = int(self.get_col_index(col_value))
        log.info('【%s】的[%s]属性从【%s】改为[%s]'%(element,col_value,df.iloc[row,col],value))
        df.iloc[row:row+1,col:col+1] = value
        print(df.iloc[row:row+1,col:col+1])
        # df.iloc[row]['by_style'] = 'username'
        self.save(df)


    '''# pandas处理（需要后期进行处理）'''
    def data_to_pandas(self):
        olddata = self.datas ()
        model,miaoshu, tag, name, by, selector,located_two = [], [], [], [], [], [], []
        index = [k for k, v in olddata.items ()]
        v = [v for k, v in olddata.items ()]
        for i in range (len (olddata)):
            model.append('')
            miaoshu.append (v[i]['discription'])
            tag.append (v[i]['tag'])
            name.append (v[i]['tag'] + ' ' + miaoshu[i][0:20])
            by.append (v[i]['by_key'])
            located_two.append('')
            selector.append ('xpath=>%s' % v[i]['xpath'])
        data = {'模块': model, '页面元素': name, '标签': tag, '关键字': index, 'by': by, 'by_style': selector, '二级定位': located_two,
            '备注': miaoshu, }
        return data
    def write_pandadata_to_xlsx(self,sheetname):
        ''''''
        path = self.path.split ('.')[0] + '(pandas).xlsx'
        data = self.data_to_pandas ()
        columns=['模块', '页面元素', '标签', '关键字', 'by', 'by_style', '二级定位','备注']
        df=pd.DataFrame(data,columns=columns)
        xlsx_writer = pd.ExcelWriter (path, engine='xlsxwriter')
        df.to_excel (xlsx_writer,sheetname, index=False)
    def rewrite_pandadata_to_xlsx(self,sheetname):
        path = self.path.split ('.')[0] + '(pandas).xlsx'
        data = self.data_to_pandas ()
        columns = ['模块', '页面元素', '标签', '关键字', 'by', 'by_style', '二级定位', '备注']
        df1 = pd.DataFrame (data, columns=columns)
        book = load_workbook (path)
        # df1 = pd.read_excel (path, index=[0])
        writer = pd.ExcelWriter (path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict ((ws.title, ws) for ws in book.worksheets)
        df1.to_excel (writer, sheetname, index=False)
        writer.save ()

        print ('成功将[%s]行element【pandas覆盖】写入%s文件！' % (len(data), path))

if __name__ == '__main__':
    path = '../exel/001.xlsx'
    s = ExcelUtil (path, 'Page')
    s2 = s.get_row_index ('重置按钮')
    s2 = s.get_row_info ('重置按钮')
    s3 = s.get_selector ('登录按钮')
    s0 = s.set_value ('登录按钮', '关键字','reset')
    s0 = s.set_value ('登录按钮', 'by','css_selector')
