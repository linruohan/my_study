# coding=utf-8
from lxml import etree
import csv,json,xlwt,xlrd,os
from pathlib import Path
from xlutils.copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter,ZipFile
import matplotlib.pyplot as plt
# from scipy.interpolate import lagrange#导入拉格朗日插值函数
class Parse_html_element():
    def __init__(self, path):
        self.path = path
    # 字典value中有中文
    def chinese_contain(self, str1):
        for i in list(str1):
            if u'\u4e00' <= i <= u'\u9fff':
                return True
        return False
    def get_chinese_from_str(self, str1):
        s = ''
        if isinstance(str1,list):
            for value in str1:
                for j in value :
                    if u'\u4e00' <= j <= u'\u9fff':
                        s += j
        else:
            for i in list((str1),):
                if u'\u4e00' <= i <= u'\u9fff':
                    s+=i
        return s
    def read(self):
        with open(self.path, 'r', encoding='utf-8')as f:
            doc = f.read()
        html = doc.replace(u'\xa9', u'').encode('utf-8', 'ignore')
        html = doc.replace(u'\xa0', u'').encode('utf-8', 'ignore')
        page = etree.HTML(html)
        return page
    def datas(self):
        page = self.read()
        htree = etree.ElementTree(page)
        data = {}
        n =0
        tag=''
        discription = ''
        by_key='xpath'
        for t in page.iter():
            try:
                tag=t.tag.__name__
                # print('t.tag.__name__==',t.tag.__name__)#标签名称
            except  AttributeError as e:
                tag = t.tag
                # print('t.tag==',t.tag)
            buyao_list=['html','head','meta','script','Comment','style','body']
            if  tag not in buyao_list:
                attrib =dict( t.attrib)
                attrib_values = [v for k, v in dict (attrib).items ()]
                attrib_keys = [k for k, v in dict (attrib).items ()]
                text = t.text
                xpath = htree.getpath (t)
                if 'id' in attrib :
                    id=attrib['id']
                    by_key='id'
                    xpath='//*[@id="%s"]'%str(id)
                if self.get_chinese_from_str(attrib_values):
                    discription = self.get_chinese_from_str (attrib_values)
                elif text and self.get_chinese_from_str(text) is not None:
                    discription=self.get_chinese_from_str(text)
                elif text and self.chinese_contain (text[:10]):
                    discription = text[:10]
                elif 'name' in attrib:
                    discription = attrib['name']
                elif 'value' in attrib:
                    discription = attrib['value']
                elif 'class' in attrib:
                    discription = attrib['class']
                else:
                    discription = u'无'
                data_key=str(n)+tag
                data[data_key] = {'index':n,'tag':tag,'discription': discription, 'by_key':by_key, 'attrib': attrib, 'xpath': xpath, 'text': text}
                by_key = 'xpath'
                n += 1
        return data
    def data_to_pandas(self):
        olddata = self.datas ()
        model,miaoshu, tag, name, by, selector,located_two = [], [], [], [], [], [], []
        index = [k for k, v in olddata.items ()]
        v = [v for k, v in olddata.items ()]
        for i in range (len (olddata)):
            model.append('')
            miaoshu.append (v[i]['discription'])
            tag.append (v[i]['tag'])
            name.append (v[i]['tag'] + ' ' + miaoshu[i][0:20])
            by.append (v[i]['by_key'])
            located_two.append('')
            selector.append ('xpath=>%s' % v[i]['xpath'])
        data = {'模块': model, '页面元素': name, '标签': tag, '关键字': index, 'by': by, 'by_style': selector, '二级定位': located_two,
            '备注': miaoshu, }
        return data
    def save_txt(self):
        olddata=self.datas()
        data1={}
        m = 0
        newfile = self.path.split ('.')[0] + '.txt'
        for k, v in olddata.items():
            miaoshu = (v['discription']).strip ()
            tag=v['tag']
            name=v['tag']+' '+miaoshu[0:20]
            by=v['by_key'].strip ()
            selector = 'xpath=>%s' % v['xpath'].strip ()
            # print(name, xpath, element)
            data = str (m) + ':' + name + '\n' + selector + '\n'
            data1[m] = {'模块': '','页面元素': name, '标签': tag, 'by': by,'by_style': selector, '二级定位': ''}
            m += 1
            with open(newfile,'a',encoding='utf-8')as f:

                f.write(data)
        print ('成功将element写入%s文件！' % newfile)
        return data1
    def save_xls(self,sheetname):
        # 创建excel工作表
        path=self.path.split ('.')[0] + '.xls'

        # 设置表头
        headers = ['模块', '页面元素', '标签', 'by', 'by_style', '二级定位']
        data = self.save_txt ()
        rb = xlwt.Workbook ()
        ws = rb.add_sheet (sheetname, cell_overwrite_ok=True)
        for i in range (len (headers)):
            ws.write (0, i, headers[i])
        n = 0
        values = [(k, v) for row, values in data.items () for k, v in values.items ()]
        for i in range (len (data)):
            if i == 15: break
            for j in range (len (headers)):
                k = [k for k, v in values]
                if k[j] == headers[j]:
                    real = data[i][k[j]]
                    ws.write (i + 1, j, real)
                j += 1
                n += 1
            i += 1
        rb.save (path)
        print ('成功将[%s]行element写入%s文件！' % (i, path))
    def re_write_xls(self,sheetname):
        # 设置表头
        path = self.path.split ('.')[0] + '.xls'
        headers = ['模块', '页面元素', '标签', 'by', 'by_style', '二级定位']
        data = self.save_txt ()
        rb = xlrd.open_workbook (path)
        sheetnames = rb.sheet_names ()
        print (sheetnames)
        wb = copy (rb)
        if sheetname in sheetnames:
            index = [i for i in sheetnames if i == sheetname][0]
            ws = wb.get_sheet (index)
        else:
            ws = wb.add_sheet (sheetname, cell_overwrite_ok=True)
        for i in range (len (headers)):
            ws.write (0, i, headers[i])
        n = 0
        values = [(k, v) for row, values in data.items () for k, v in values.items ()]
        for i in range (len (data)):
            if i == 15: break
            for j in range (len (headers)):
                k = [k for k, v in values]
                if k[j] == headers[j]:
                    real = data[i][k[j]]
                    ws.write (i + 1, j, real)
                j += 1
                n += 1
            i += 1
        wb.save (path)
        print ('成功将[%s]行element【覆盖】写入%s文件！' % (i, path))
    def write_pandadata_to_xlsx(self,sheetname):
        ''''''
        path = self.path.split ('.')[0] + '(pandas).xlsx'
        data = self.data_to_pandas ()
        columns=['模块', '页面元素', '标签', '关键字', 'by', 'by_style', '二级定位','备注']
        df=pd.DataFrame(data,columns=columns)
        xlsx_writer = pd.ExcelWriter (path, engine='xlsxwriter')
        df.to_excel (xlsx_writer,sheetname, index=False)
    def rewrite_pandadata_to_xlsx(self,sheetname):
        path = self.path.split ('.')[0] + '(pandas).xlsx'
        data = self.data_to_pandas ()
        columns = ['模块', '页面元素', '标签', '关键字', 'by', 'by_style', '二级定位', '备注']
        df1 = pd.DataFrame (data, columns=columns)
        book = load_workbook (path)
        # df1 = pd.read_excel (path, index=[0])
        writer = pd.ExcelWriter (path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict ((ws.title, ws) for ws in book.worksheets)
        df1.to_excel (writer, sheetname, index=False)
        writer.save ()

        print ('成功将[%s]行element【pandas覆盖】写入%s文件！' % (len(data), path))
    def save_openpyxl(self,sheetname):
        '''不能覆盖写'''
        # 新建一个workbook
        wb = Workbook ()
        # ws1 = wb.create_sheet ()  # 默认插在最后
        # ws2 = wb.create_sheet (0)  # 插在开头
        sheets = wb.worksheets
        sheetnames=wb.sheetnames
        for name in sheetnames:
            if name==sheetname:
                ws=wb.get_sheet_by_name(sheetname)
            else:
                ws = wb.create_sheet (index=0,title = sheetname)
        # archive=ZipFile('001.txt', 'w', 8, allowZip64=True)
        # ew = ExcelWriter (workbook=wb,archive = archive)
        path = self.path.split ('.')[0] + '.xlsx'
        data = self.save_txt ()
        # 设置表头
        headers = ['模块', '页面元素', '标签', 'by', 'by_style', '二级定位']
        for i in range (len (headers)):
            s=ws.cell (row=1, column=i+1)
            s.value=headers[i]
        wb.save (path)
        n = 1
        values = [(k, v) for row, values in data.items () for k, v in values.items ()]
        for i in range (len (data)):
            # if i == 15: break
            for j in range (len (headers)):
                k = [k for k, v in values]
                v = [v for k, v in values]
                if k[j] == headers[j]:
                    real = data[i][k[j]]
                    print (real)
                    s1=ws.cell (row=i + 1,column=j+1)
                    s1.value=real
                j += 1
                n += 1
            i += 1
        wb.save (path)
        # ew.save(path)(使用ExcelWriter保存)
        print ('使用openpyxls成功将[%s]行element写入%s文件！' % (i, path))



if __name__ == '__main__':
    path = '卡口设备状态.html'
    s=Parse_html_element(path)
    # s.save_txt()
    # s.save_xls('sheet12sad')
    # s.re_write_xls('123')
    # s.write_pandadata_to_xlsx('卡口设备状态')
    s.rewrite_pandadata_to_xlsx('卡口设备状态01')