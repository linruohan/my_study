import pandas as pd
import numpy as np
import pandas
import xlrd,xlwt
from openpyxl import load_workbook

path = '001.xlsx'
def read():
    df1 = pd.read_excel(path, 'PageElements', index_col=None, na_values=['9999'])
    # print(df)
    print(df1['关键字'][0])
    return df1
def save(path, a, sheetname):
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    a.to_excel(writer, sheetname, index=False)
    writer.save()


df = read()
df['关键字'][0]= 'yueer'
print("*" * 20)
save(path, df, 'PageElements')
read()
'''
read_excel(io, sheetname=0, header=0,
skiprows=None, skip_footer=0,
index_col=None,names=None,
parse_cols=None, parse_dates=False,
date_parser=None,
na_values=None,thousands=None,
convert_float=True, has_index_names=None,
 converters=None,dtype=None,
 true_values=None,
 false_values=None, engine=None,
 squeeze=False, **kwds)
io : string, path object ; excel 路径。
sheetname : string, int, mixed list of strings/ints,
        or None, default 0 返回多表使用sheetname=[0,1],
        若sheetname=None是返回全表 注意：int/string 返回的是dataframe，
        而none和list返回的是dict of dataframe
header : int, list of ints, default 0 指定列名行，
        默认0，即取第一行，数据为列名行以下的数据
        若数据不含列名，则设定 header = None
skiprows : list-like,Rows to skip at the beginning，
        省略指定行数的数据
skip_footer : int,default 0, 省略从尾部数的int行数据
index_col : int, list of ints, default None指定列为索引列，
        也可以使用u”strings”
names : array-like, default None, 指定列的名字。
'''
# *******************************************************
'''
DataFrame.to_excel（excel_writer，sheet_name ='Sheet1'，
na_rep =''，float_format = None，
columns = None，header = True，
index = True，index_label = None，
startrow = 0，startcol = 0，engine = None，
merge_cells = True，encoding = None，
inf_rep ='inf'，verbose = True，
freeze_panes = None ）
将DataFrame写入Excel工作表

参数：
excel_writer：字符串或ExcelWriter对象
        文件路径或现有的ExcelWriter
sheet_name：字符串，默认'Sheet1'
        将包含DataFrame的工作表名称
na_rep：字符串，默认为''
        缺少数据表示
float_format：字符串，默认无
        格式化浮点数的字符串
列：序列，可选
        要编写的列
标题：布尔值或字符串列表，默认为True 写出列名。如果给出了字符串列表，则假定它是列名称的别名
索引：布尔值，默认为True 写行名（索引）
index_label：字符串或序列，默认无
        如果需要，索引列的列标签。如果给出无，并且 标题和索引为真，则使用索引名称。如果DataFrame使用MultiIndex，则应该给出一个序列。
startrow：左上角的单元格行来转储数据帧
startcol：左上角的单元格列转储数据帧
引擎：字符串，默认无
        写引擎使用-您也可以通过选项设置此 io.excel.xlsx.writer，io.excel.xls.writer和 io.excel.xlsm.writer。
merge_cells：布尔值，默认为True
        将MultiIndex和Hierarchical行编写为合并单元格。
编码：字符串，默认无
        编码生成的excel文件。只有xlwt需要，其他编写者本地支持unicode。
inf_rep：字符串，默认'inf'
        无限的表示（在Excel中无限制的表示）
freeze_panes：整数（长度为2）的元组，默认为无
        指定要冻结的基于1的最底部行和最右边的列
        新版本0.20.0。'''
